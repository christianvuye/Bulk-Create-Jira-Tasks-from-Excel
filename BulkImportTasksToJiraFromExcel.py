#import libraries
from openpyxl import load_workbook
from jira import JIRA
from enum import Enum
from datetime import datetime
import os
import xlsxwriter

#connect to Jira and open Excel worksheet for reading
# replace 'https://jira.atlassian.com' with the Jira server you are creating the bulk tasks for 
# replace 'username' and 'password' with login details from a user who has the appropriate rights to create tasks
print("Connecting and signing user into Jira...")
jira = JIRA('https://jira.atlassian.com',
                     basic_auth=('username', 'password'))
print("Succesfully connected and signed in to Jira as user.")

print("Reading Excel file...")
wb = load_workbook(filename = 'Excel_Template.xlsx', read_only = True)
print("Excel file succesfully loaded.")

print("Loading Excel sheet...")
ws = wb['ExcelToJira']
print("Excel sheet loaded.")

#constants
LIMIT_OF_SINGLE_DIGIT_NUMBERS = 9
FIRST_ROW = 2
FIRST_COLUMN = 1
LAST_ROW = ws.max_row+1
LAST_COLUMN = 12

#functions
def printAmountofIssuesCreated():
	print("DONE\t Total new tssues created: " + str(i-1))
	exit()

def promp():
	print("Press Any Key To Exit.")
	input()
	printAmountofIssuesCreated()

def getCellValue(worksheet,  r, c):
	return worksheet.cell(row=r, column=c).value

def EndOfSheetReached():
	print ("The end of the sheet has been reached.")
	promp()

def AddAttachment(list_of_files_with_particular_extension):
    n = 1
    for every_file_with_particular_extension in list_of_files_with_particular_extension:
        print(every_file_with_particular_extension)
        every_file_with_particular_extension_split = every_file_with_particular_extension.split("_")
        print(every_file_with_particular_extension_split)
        every_file_with_particular_extension_split_filename = every_file_with_particular_extension_split[0]
        print(every_file_with_particular_extension_split_filename)
        every_file_with_particular_extension_split_fileextension = every_file_with_particular_extension_split[1].lower()
        print(every_file_with_particular_extension_split_fileextension)
        print(str(column_list[eColumn.SUMMARY.value]).strip())
        if every_file_with_particular_extension_split_filename == str(column_list[eColumn.SUMMARY.value].strip()):
            print("First image found in list which is equal to summary string.")
            print(every_file_with_particular_extension_split_fileextension.replace(str(l),""))
            n = int(every_file_with_particular_extension_split_fileextension.replace(str(l),""))
            print("Gave n the value of %s." % (str(n)))
            if n <= LIMIT_OF_SINGLE_DIGIT_NUMBERS:
                attachment_name = jira.add_attachment(issue,str((str(column_list[eColumn.SUMMARY.value])).strip() + "_" + str(n) + str(l)))
                print("Attachment %s has been added." % str(attachment_name))
            else:
                attachment_name = jira.add_attachment(issue,str((str(column_list[eColumn.SUMMARY.value])).strip() + "_" + str(n) + str(l)))
                print("Attachment %s has been added." % str(attachment_name))
            

def checkFileFormatsAndGroupIt(file_format):
	#returns all the files in the current directory and stores it in "files"
	files = os.listdir() 
	#prints all the files in the current directory
	print (files) 
	#declare empty list to which elements will later be added
	file_list_grouped_by_extension = [] 
	for each_file in files:
		#prints complete filename, including extension 
		print(each_file)
		#prints 'filename','extension' as tuple  
		print(os.path.splitext(each_file)) 
		#prints the type, i.e. "tuple"
		print(type(os.path.splitext(each_file)))
		#access and print the second element in the tuple, i.e. the file extension, as a string 
		print(os.path.splitext(each_file)[1]) 
		#prints the type of the second element in the tuple, ie. string
		print(type(os.path.splitext(each_file)[1])) 
		#if the fileformat of the file is equal to the file format given in the argument of the function,
		if os.path.splitext(each_file)[1].lower() == file_format.lower():
			#add the complete filename and extension string to the list 
			file_list_grouped_by_extension.append(each_file) 
			print("%s has been added to the list" % (each_file))
	#and return that list
	return file_list_grouped_by_extension 

#classes	
class eColumn(Enum):
	PROJECT = 0
	ISSUE_TYPE = 1
	SUMMARY = 2
	PARENT_WATCHER = 3
	PRIORITY = 4
	DUE_DATE = 5
	EXTERNAL_BID = 6
	ASSIGNEE = 7
	COMPONENT = 8
	DESCRIPTION = 9
	FIX_VERSION = 10


#declare and initialize variables 
i = 0
j = 0
k = 0
l = 0
m = 0	

file_formats = [
	".jpg",
	".jpeg",
	".doc",
	".docx",
	".svg",
	".bmp",
]

#start script
for i in range (FIRST_ROW, LAST_ROW):
	column_list = []
	
	for j in range(FIRST_COLUMN,LAST_COLUMN):
		column_list.append(getCellValue(ws,i,j))
	
	if column_list[eColumn.PROJECT.value] == None and column_list[eColumn.ISSUE_TYPE.value] == None and column_list[eColumn.SUMMARY.value] == None:
		EndOfSheetReached()
	elif None in column_list:
		print("An empty cell has been encountered on row %s and column %s. Please fill in the missing data and run the program again." % (i,xlsxwriter.utility.xl_col_to_name(column_list.index(None))))
		promp()
	
	for k in range(eColumn.PROJECT.value,eColumn.FIX_VERSION.value+1):
		print(column_list[k]) 

	project_components = jira.project_components(column_list[eColumn.PROJECT.value])
	comp_string = []
	for comp in project_components:
		comp_string.append(str(comp))
	if column_list[eColumn.COMPONENT.value] not in comp_string:
		print ("%s is not an existing component." % column_list[eColumn.COMPONENT.value])
		new_component = jira.create_component(column_list[eColumn.COMPONENT.value], jira.project(column_list[eColumn.PROJECT.value]), \
		                description=None, leadUserName=None, assigneeType=None, isAssigneeTypeValid=False)
		print ("Created a new component called %s." % column_list[eColumn.COMPONENT.value])
	else:
		print ("%s is an existing component." % column_list[eColumn.COMPONENT.value])	

	project_fixversions = jira.project_versions(column_list[eColumn.PROJECT.value])
	fixv_string = []
	for fixv in project_fixversions:
		fixv_string.append(str(fixv))
	if column_list[eColumn.FIX_VERSION.value] not in fixv_string:
		print ("%s is not an an existing fixversion." % column_list[eColumn.FIX_VERSION.value])
		new_fixversion = jira.create_version(column_list[eColumn.FIX_VERSION.value], jira.project(column_list[eColumn.PROJECT.value]),\
		                 description=None, releaseDate=None, startDate=None, archived=False, released=False)
		print ("Created a new fixversion called %s." % column_list[eColumn.FIX_VERSION.value])
	else:
		print ("%s is an existing fixversion." % column_list[eColumn.FIX_VERSION.value])
	
	issue = jira.create_issue(project = column_list[eColumn.PROJECT.value],issuetype=column_list[eColumn.ISSUE_TYPE.value],summary = column_list[eColumn.SUMMARY.value],\
							  customfield_13700 = { "name": column_list[eColumn.PARENT_WATCHER.value]}, priority = {'name': column_list[eColumn.PRIORITY.value]}, \
							  duedate=str(column_list[eColumn.DUE_DATE.value].strftime('%Y-%m-%d')) +"T00:00:00.000-0500", customfield_12501 = column_list[eColumn.EXTERNAL_BID.value], \
							  components = [{"name" : column_list[eColumn.COMPONENT.value]}], description=column_list[eColumn.DESCRIPTION.value], \
							  fixVersions  = [{"name" : column_list[eColumn.FIX_VERSION.value]}])
	jira.assign_issue(issue,column_list[eColumn.ASSIGNEE.value])
	
	for l in file_formats:
		file_list_grouped_by_extension_1 = checkFileFormatsAndGroupIt(l)
		print(file_list_grouped_by_extension_1)
		AddAttachment(file_list_grouped_by_extension_1)
	
	print("Successfuly Created: "+ str(column_list[eColumn.ISSUE_TYPE.value]) + " " + str(issue) + " " + str(column_list[eColumn.SUMMARY.value]))	

promp()